from CA_Fields import CA_Fields
import openpyxl
import datetime


def process_xlsx(file, submitter_id='0049', auditor_id='000', batch_num='001'):
    wb = openpyxl.load_workbook(file)
    sheet = wb["ACCESS"]

    # PREPARE STATIC DATA

    req_rec_type = 5
    req_rec_num = ' ' * 7
    head_rec_type = 1
    #submitter_id = '0049'
    #auditor_id = '0000'
    #batch_num = '001'


    #SSN = '0' * 9
    #ID = ' ' * 14
    now = datetime.datetime.now()
    year = now.year
    month = now.month
    day = now.day

    # ==================================================================
    # should curr_date_str be used for both File creation date in header 
    # and also rec_creation_date in rows ?????
    #
    curr_date_str = f'{year:4d}{month:02d}{day:02d}'
    # ==================================================================
    # PROCESS RECORDS

    req_rows = ''

    # Track earliest & latest service dates in file and check each row as processing continues
    #   these max/min service dates will be used in building header row for request
    earliest_service_date = now
    latest_service_date = datetime.datetime(1900,1,1)

    for curr_row in range(2,sheet.max_row+1):
            row_str = f'{req_rec_type}{curr_date_str}{req_rec_num}'
            SSN = sheet.cell(row=curr_row, column=CA_Fields["MPIPTSSN"]).value
            if SSN == None:
                    row_str += '0' * 9
            else:
                    row_str += SSN.ljust(9,'0')
            #row_str += sheet.cell(row=curr_row, column=CA_Fields["MPIPTMCALID"]).value.ljust(14,0)
            CALID = sheet.cell(row=curr_row, column=CA_Fields["MPIPTMCALID"]).value
            if CALID == None:
                    row_str += ('0' * 9).ljust(14)
            else:
                    row_str += CALID.ljust(14)
            

            row_str += f'{sheet.cell(row=curr_row, column=CA_Fields["MPIPTLASTNAME"]).value:20s}'
            row_str += f'{sheet.cell(row=curr_row, column=CA_Fields["MPIPTFIRSTNAME"]).value:15s}'
            row_str += FormatDateStr(sheet.cell(row=curr_row, column=CA_Fields["MPIPTDOB"]).value)
            row_str += f'{sheet.cell(row=curr_row, column=CA_Fields["MPIPTGENDER"]).value}'
            doa = CreateDate(sheet.cell(row=curr_row, column=CA_Fields["MPIPTDOA"]).value)
            dod = CreateDate(sheet.cell(row=curr_row, column=CA_Fields["MPIPTDOD"]).value)

            # Track earliest service date in file and check each row as processing continues
            if doa < earliest_service_date:
                    earliest_service_date = doa

            # Track latest service date in file and check each row as processing continues
            if dod > latest_service_date:
                    latest_service_date = dod

            curr_date = doa
            while curr_date <= dod:
                    #print(row_str + f'{curr_date.year:4d}{curr_date.month:02d}')
                    req_rows += row_str + f'{curr_date.year:4d}{curr_date.month:02d}' + '\n'
                    curr_date = IncrementMonth(curr_date)


    # FINISHED CREATING REQ ROWS

    # CREATE HEADER
    provider_id = sheet.cell(row=2, column=CA_Fields["MPIPRVDR"]).value
    head_str = str(head_rec_type) + curr_date_str + submitter_id
    head_str += batch_num + provider_id + auditor_id
    head_str += FormatDate(earliest_service_date) + FormatDate(latest_service_date) + '\n'

    #print(head_str + req_rows)	

    out_file = open('out.txt','w')
    out_file.write(head_str + req_rows)
    out_file.close()	

# UTILITY FUNCTIONS

def FormatDateStr(date_str):
	month, day, year = date_str.split('/')
	return f'{year:4s}{month:2s}{day:2s}'

def FormatDate(d):
	return f'{d.year:04d}{d.month:02d}{d.day:02d}'

def CreateDate(date_str):
	month, day, year = date_str.split('/')
	return datetime.datetime(int(year), int(month),int(day))

def IncrementMonth(d):
	if d.month == 12:
		d2 = datetime.datetime(d.year+1,1,1)
	else:
		d2 = datetime.datetime(d.year,d.month+1,1)

	return d2
